"""
순수 비즈니스 로직 — Streamlit 의존 없음
견적서 Excel 생성, 단가표/카탈로그 로드
"""
import os, io, re
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from datetime import date
from dotenv import load_dotenv

load_dotenv()

# ─── 경로 상수 ───────────────────────────────────────────────────
DATA_DIR = os.environ.get("ANSYS_DATA_DIR", "")
PRICELIST_FILE = os.path.join(DATA_DIR, "#. ANSYS 2026 Pricelist_최종.xlsx")
BU_FILES = {
    "MBU":          "01. Commercial MBU_영업공통_견적서.xlsx",
    "FBU":          "02. Commercial FBU_영업공통_견적서.xlsx",
    "EBU":          "03. Commercial EBU_영업공통_견적서.xlsx",
    "SBU":          "04. Commercial SBU_영업공통_견적서.xlsx",
    "CPBU+DBU+Matbu": "05. Commercial CPBU+DBU+Matbu_영업공통_견적서.xlsx",
    "Startup":      "06. Startup_임대견적서.xlsx",
}
HIDDEN_SHEETS   = {"제품군", "단가", "담당자"}
FORMULA_SHEETS  = {"단가", "담당자", "제품군"}


# ─── 단가표 로드 ────────────────────────────────────────────────
def load_pricelist() -> pd.DataFrame:
    wb = openpyxl.load_workbook(PRICELIST_FILE, data_only=True)
    ws = wb["UPLIFT_26년"]
    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        name = row[0]
        if not name or not isinstance(name, str) or not name.strip():
            continue
        name = name.strip()
        if name.startswith(("Note", "Pricing", "Discount", "Product")):
            continue
        try:
            perpetual = int(row[1]) if row[1] and row[1] not in ("NA", "N/A") else 0
            tecs      = int(row[3]) if row[3] and row[3] not in ("NA", "N/A") else 0
            lease     = int(row[4]) if row[4] and row[4] not in ("NA", "N/A") else 0
        except (ValueError, TypeError):
            continue
        if perpetual > 0 or lease > 0:
            rows.append({"제품명": name, "구매(영구)": perpetual,
                         "TECS(유지보수)": tecs, "임대(연간)": lease})
    return pd.DataFrame(rows)


# ─── 제품 카탈로그 구축 ──────────────────────────────────────────
def build_product_catalog() -> dict:
    """
    반환: {product_name: {"구매": {bu,sheet,desc} | None, "임대": {...} | None}}
    """
    catalog = {}
    for bu_key, fname in BU_FILES.items():
        path = os.path.join(DATA_DIR, fname)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        for sname in wb.sheetnames:
            if sname in HIDDEN_SHEETS:
                continue
            ws = wb[sname]
            lic_text = str(ws["M6"].value or "").strip()
            if not lic_text:
                continue
            lic_cat = "임대" if ("Annual" in lic_text or "Maintenance" in lic_text) else "구매"

            for r in range(18, 65):
                b_val = str(ws.cell(r, 2).value or "").strip()
                c_val = str(ws.cell(r, 3).value or "").strip()
                if re.match(r'^\d+\.$', b_val) and c_val and len(c_val) > 3:
                    pname = c_val
                    if pname not in catalog:
                        catalog[pname] = {"구매": None, "임대": None}
                    if catalog[pname][lic_cat] is not None:
                        continue
                    desc = []
                    for dr in range(r + 1, r + 35):
                        b2 = str(ws.cell(dr, 2).value or "").strip()
                        c2 = str(ws.cell(dr, 3).value or "").strip()
                        d2 = str(ws.cell(dr, 4).value or "").strip()
                        a2 = ws.cell(dr, 1).value
                        if b2.endswith(".") and b2 != ".":
                            break
                        if a2 is not None and isinstance(a2, (int, float)) and float(a2) >= 2:
                            break
                        line = (d2 or c2).strip()
                        if line and "●" not in line:
                            desc.append(line)
                        if len(desc) >= 15:
                            break
                    catalog[pname][lic_cat] = {"bu": bu_key, "sheet": sname, "desc": desc}
    return catalog


# ─── 단가 조회 ──────────────────────────────────────────────────
def get_price(df: pd.DataFrame, name: str, lic: str) -> int:
    col = "구매(영구)" if lic == "구매" else "임대(연간)"
    row = df[df["제품명"] == name]
    if row.empty:
        row = df[df["제품명"].str.contains(
            name.split()[1] if len(name.split()) > 1 else name,
            case=False, na=False)]
    if row.empty:
        return 0
    v = row.iloc[0][col]
    return int(v) if v else 0


# ─── 출력 Excel 파싱 (검증용) ───────────────────────────────────
def read_products_from_excel(buf: io.BytesIO) -> list[dict]:
    """
    생성된 Excel에서 제품 행을 읽어 반환.
    반환: [{"name": str, "qty": int, "price": int}, ...]
    """
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb.active
    products = []
    for r in range(18, 120):
        b_val = str(ws.cell(r, 2).value or "").strip()
        c_val = str(ws.cell(r, 3).value or "").strip()
        i_val = ws.cell(r, 9).value
        j_val = ws.cell(r, 10).value
        if (re.match(r'^\d+\.$', b_val) and c_val
                and i_val is not None
                and isinstance(i_val, (int, float))
                and float(i_val) > 0):
            products.append({
                "name":  c_val,
                "qty":   int(i_val),
                "price": int(j_val) if j_val else 0,
            })
    return products


# ─── 템플릿 인덱스 + 자동 선택 ──────────────────────────────────
def build_template_index() -> list[dict]:
    """
    모든 BU 파일을 스캔해 사용 가능한 템플릿 목록을 반환.
    반환: [{"bu": str, "sheet": str, "slot_count": int, "lic_cat": str}, ...]
    slot_count = price_rows 개수, lic_cat = "구매" | "임대"
    정렬: slot_count 오름차순
    """
    index = []
    for bu_key, fname in BU_FILES.items():
        path = os.path.join(DATA_DIR, fname)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        for sname in wb.sheetnames:
            if sname in HIDDEN_SHEETS:
                continue
            if wb[sname].sheet_state != "visible":
                continue
            ws = wb[sname]
            lic_text = str(ws["M6"].value or "").strip()
            lic_cat = "임대" if ("Annual" in lic_text or "Maintenance" in lic_text) else "구매"
            price_rows = []
            for r in range(18, 120):
                b_val = str(ws.cell(r, 2).value or "").strip()
                i_val = ws.cell(r, 9).value
                if (re.match(r'^\d+\.$', b_val)
                        and i_val is not None
                        and isinstance(i_val, (int, float))):
                    price_rows.append(r)
            index.append({
                "bu": bu_key,
                "sheet": sname,
                "slot_count": len(price_rows),
                "lic_cat": lic_cat,
            })
    index.sort(key=lambda x: x["slot_count"])
    return index


def select_template(index: list[dict], n_items: int, lic_key: str) -> tuple[str, str]:
    """
    아이템 수를 수용 가능한 슬롯을 가진 가장 작은 적합 템플릿 반환.
    lic_key: "구매" | "임대"
    반환: (bu_key, sheet_name)
    """
    # 1. 라이선스 유형 일치 + 슬롯 충분한 첫 번째
    for t in index:
        if t["lic_cat"] == lic_key and t["slot_count"] >= n_items:
            return t["bu"], t["sheet"]
    # 2. 라이선스 무관, 슬롯 충분한 첫 번째
    for t in index:
        if t["slot_count"] >= n_items:
            return t["bu"], t["sheet"]
    # 3. Fallback: 슬롯이 가장 큰 템플릿
    best = max(index, key=lambda x: x["slot_count"])
    return best["bu"], best["sheet"]


# ─── Excel 생성 ─────────────────────────────────────────────────
def generate_excel(bu_key: str, sheet_name: str, info: dict, items: list,
                   disc_pct: int = 0) -> io.BytesIO:
    """
    items:    [{"name": str, "qty": int, "price": int, "desc": [str]}]
    info:     {"customer", "contact", "tel", "our_name", "our_tel", "our_email", "issue_date"}
    disc_pct: 할인율 0~100 (기본값 0)
    """
    path = os.path.join(DATA_DIR, BU_FILES[bu_key])
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else None
    if ws is None:
        visible = [s for s in wb.sheetnames
                   if wb[s].sheet_state == "visible" and s not in HIDDEN_SHEETS]
        ws = wb[visible[0]] if visible else wb[wb.sheetnames[0]]

    # ── outline(+/- 버튼) 제거 ─────────────────────────────────
    for rd in ws.row_dimensions.values():
        rd.outlineLevel = 0
        rd.hidden = False          # collapse된 rows 47-56 unhide
    ws.sheet_format.outlineLevelRow = 0

    # ── 고객 정보 ──────────────────────────────────────────────
    ws["B8"]  = info.get("customer", "") + " 貴中"
    ws["D10"] = info.get("contact", "")
    ws["D11"] = info.get("tel", "")
    ws["K9"]  = info["issue_date"].strftime("%Y. %m. %d")
    ws["K10"] = info.get("our_name", "")
    ws["K11"] = info.get("our_tel", "")
    ws["K12"] = info.get("our_email", "")

    def _write(row, col, val):
        """MergedCell의 경우 master cell을 찾아서 기록한다."""
        cell = ws.cell(row, col)
        # MergedCell이면 해당 merge 범위의 좌상단(master)을 찾아 기록
        for rng in ws.merged_cells.ranges:
            if (row, col) in [(r, c) for r in range(rng.min_row, rng.max_row + 1)
                              for c in range(rng.min_col, rng.max_col + 1)]:
                master = ws.cell(rng.min_row, rng.min_col)
                master.value = val
                return
        cell.value = val

    # ── 템플릿의 모든 가격 행 탐색 ─────────────────────────────
    # 기준: B열 = "숫자." 패턴  AND  I열(수량) = 숫자
    price_rows: list[int] = []
    for r in range(18, 120):
        b_val = str(ws.cell(r, 2).value or "").strip()
        i_val = ws.cell(r, 9).value
        if (re.match(r'^\d+\.$', b_val)
                and i_val is not None
                and isinstance(i_val, (int, float))):
            price_rows.append(r)

    # 가격 행이 없으면 I열에 숫자가 있는 첫 행을 폴백으로 사용
    if not price_rows:
        for r in range(18, 65):
            i_val = ws.cell(r, 9).value
            if i_val is not None and isinstance(i_val, (int, float)) and float(i_val) > 0:
                price_rows = [r]
                break
    if not price_rows:
        price_rows = [20]

    # ── 기존 데이터 초기화 (모든 가격행의 B~K열) ──────────────
    for r in price_rows:
        _write(r, 2, None)   # B = 번호
        _write(r, 3, None)   # C = 제품명
        _write(r, 9, None)   # I = 수량
        _write(r, 10, None)  # J = 단가
        _write(r, 11, None)  # K = 금액

    # ── 아이템 쓰기 + 설명 행 처리 ─────────────────────────────
    # 금액: Pricelist에서 가져온 item["price"] 사용
    # 설명: 카탈로그가 각 제품의 견적서 시트에서 읽은 desc 사용
    #       없으면 공란 처리
    DESC_MAX = 15

    def _clear_desc_zone(r_s: int, r_e: int):
        """설명 구역 A~E열 전체 초기화"""
        for dr in range(r_s, r_e + 1):
            _write(dr, 1, None)   # A (outline 번호 잔여 제거)
            _write(dr, 2, None)   # B
            _write(dr, 3, None)   # C
            _write(dr, 4, None)   # D
            _write(dr, 5, None)   # E (날짜 등 잔여 제거)
            ws.row_dimensions[dr].height = None   # auto 높이로 통일

    def _write_desc(r_s: int, r_e: int, desc: list):
        """초기화된 구역에 desc 텍스트를 C열에 순서대로 기록 (좌측 정렬)"""
        d_idx = 0
        for dr in range(r_s, r_e + 1):
            if d_idx >= len(desc):
                break
            if desc[d_idx].strip():
                _write(dr, 3, desc[d_idx])
                ws.cell(dr, 3).alignment = Alignment(horizontal='left', wrap_text=False)
            d_idx += 1

    for i, item in enumerate(items):
        if i < len(price_rows):
            r = price_rows[i]
        else:
            r = price_rows[-1] + (i - len(price_rows) + 1)

        # 제품 행
        _write(r, 2, f"{i + 1}.")
        _write(r, 3, item["name"])
        _write(r, 9, item["qty"])
        _write(r, 10, item["price"])
        _write(r, 11, item["qty"] * item["price"])

        # 설명 구역
        # - 클리어: 다음 가격 행 직전까지 전부 (D열 잔재 포함)
        # - 쓰기:   DESC_MAX 행까지만 (catalog 수집 한도와 동일)
        r_start      = r + 1
        if i + 1 < len(price_rows):
            r_clear_end = price_rows[i + 1] - 1
        else:
            r_clear_end = r + DESC_MAX          # 기본값
            for _dr in range(r + 1, r + DESC_MAX + 10):
                if ws.cell(_dr, 9).value is not None:   # I열 경계 = Total Amount 등
                    r_clear_end = _dr - 1
                    break
        r_write_end  = min(r_clear_end, r + DESC_MAX)

        _clear_desc_zone(r_start, r_clear_end)
        _write_desc(r_start, r_write_end, item.get("desc") or [])

    # ── 여분 가격 행 초기화 (번호/제품명/수량/단가/금액 모두) ──
    for ei, extra_r in enumerate(price_rows[len(items):], start=len(items)):
        _write(extra_r, 2, "")    # B = 번호
        _write(extra_r, 3, "")    # C = 제품명
        _write(extra_r, 9, None)  # I = 수량
        _write(extra_r, 10, None) # J = 단가
        _write(extra_r, 11, None) # K = 금액
        nxt_idx = ei + 1
        if nxt_idx < len(price_rows):
            r_clear_end = price_rows[nxt_idx] - 1
        else:
            r_clear_end = extra_r + DESC_MAX
            for _dr in range(extra_r + 1, extra_r + DESC_MAX + 10):
                if ws.cell(_dr, 9).value is not None:
                    r_clear_end = _dr - 1
                    break
        _clear_desc_zone(extra_r + 1, r_clear_end)

    # ── 할인 적용: Special D/C + Revised Price 행 기록 ───────────
    if disc_pct > 0:
        subtotal = sum(it["qty"] * it["price"] for it in items)
        disc_amt = int(subtotal * disc_pct / 100)
        final    = subtotal - disc_amt
        for r in range(50, 85):
            i_val = str(ws.cell(r, 9).value or "")
            if "Total Amount" in i_val:
                # I·K열 병합(I57:J59, K57:K59) 해제 → 개별 셀로 분리
                # A57:H59 병합(cols 1-8)은 조건에 걸리지 않으므로 보존됨
                for rng in list(ws.merged_cells.ranges):
                    if (rng.min_row <= r + 2 and rng.max_row >= r + 1
                            and rng.max_col >= 9 and rng.min_col <= 11):
                        ws.unmerge_cells(str(rng))
                # unmerge 후 직접 기록 (_write 우회: master 탐색 불필요)
                ws.cell(r + 1, 9).value  = "Special D/C :"
                ws.cell(r + 1, 11).value = disc_amt
                ws.cell(r + 2, 9).value  = "Revised Price :"
                ws.cell(r + 2, 11).value = final
                break

    # ── 불필요 시트 삭제, 수식 참조 시트는 숨김 유지 ───────────
    active_sheet = ws.title
    for s in [x for x in wb.sheetnames if x != active_sheet and x not in FORMULA_SHEETS]:
        del wb[s]
    for s in FORMULA_SHEETS:
        if s in wb.sheetnames:
            wb[s].sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
