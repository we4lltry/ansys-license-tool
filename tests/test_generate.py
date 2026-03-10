#!/usr/bin/env python3
"""
견적서 자동 검증 테스트
====================================
인풋(제품/수량/단가) 대비 아웃풋 Excel이 일치하는지 자동으로 검증합니다.
Streamlit 없이 단독 실행 가능.

사용법:
    cd ansys-license-tool
    python tests/test_generate.py
"""
import sys, os, io

# 프로젝트 루트를 경로에 추가
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from datetime import date
from utils.excel_gen import (
    generate_excel,
    load_pricelist,
    build_product_catalog,
    get_price,
    read_products_from_excel,
    BU_FILES, DATA_DIR,
)

# ─── 공통 더미 발행자 정보 ───────────────────────────────────────
DUMMY_INFO = {
    "customer":   "테스트컴퍼니",
    "contact":    "홍길동 부장",
    "tel":        "02-1234-5678",
    "our_name":   "김담당",
    "our_tel":    "010-9999-0000",
    "our_email":  "test@tsne.co.kr",
    "issue_date": date(2026, 3, 4),
}


# ─── 검증 헬퍼 ──────────────────────────────────────────────────
class TestResult:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.errors: list[str] = []

    def ok(self, msg: str):
        self.passed += 1
        print(f"    [PASS] {msg}")

    def fail(self, msg: str):
        self.failed += 1
        self.errors.append(msg)
        print(f"    [FAIL] {msg}")

    def summary(self) -> bool:
        total = self.passed + self.failed
        print(f"\n{'='*55}")
        print(f"결과: {self.passed}/{total} 통과", end="")
        if self.failed:
            print(f"  ({self.failed}개 실패)")
        else:
            print("  — 모두 통과!")
        print('='*55)
        return self.failed == 0


def check(result: TestResult, label: str, condition: bool, detail: str = ""):
    if condition:
        result.ok(label + (f" ({detail})" if detail else ""))
    else:
        result.fail(label + (f" → {detail}" if detail else ""))


def verify_output(result: TestResult, label: str,
                  input_items: list[dict], buf: io.BytesIO):
    """출력 Excel에서 제품 행을 읽어 인풋과 비교"""
    buf.seek(0)
    output = read_products_from_excel(buf)

    in_names  = {it["name"] for it in input_items}
    out_names = {p["name"]  for p in output}

    missing = in_names - out_names
    extra   = out_names - in_names

    print(f"\n  [{label}]")
    print(f"  입력 {len(input_items)}개: {[it['name'] for it in input_items]}")
    print(f"  출력 {len(output)}개:  {[p['name'] for p in output]}")

    check(result, "제품 수 일치",
          len(output) == len(input_items),
          f"기대={len(input_items)}, 실제={len(output)}")

    check(result, "누락 없음",
          not missing,
          f"누락={missing}" if missing else "")

    check(result, "여분 없음",
          not extra,
          f"여분={extra}" if extra else "")

    # 수량/단가 검증
    out_map = {p["name"]: p for p in output}
    for it in input_items:
        if it["name"] not in out_map:
            continue
        p = out_map[it["name"]]
        check(result, f"수량 일치 [{it['name'][:30]}]",
              p["qty"] == it["qty"],
              f"기대={it['qty']}, 실제={p['qty']}")
        check(result, f"단가 일치 [{it['name'][:30]}]",
              p["price"] == it["price"],
              f"기대={it['price']:,}, 실제={p['price']:,}")


# ─── 데이터 로드 ────────────────────────────────────────────────
def load_data():
    print("단가표 및 카탈로그 로딩 중...", end=" ", flush=True)
    df = load_pricelist()
    cat = build_product_catalog()
    print(f"완료 (제품 {len(df)}개, 카탈로그 {len(cat)}개)\n")
    return df, cat


def pick_products(cat: dict, df, lic_key: str, n: int = 2) -> list[dict]:
    """카탈로그에서 지정 라이선스 유형 제품 n개 반환"""
    chosen = []
    for name, entry in cat.items():
        if entry.get(lic_key) is None:
            continue
        price = get_price(df, name, lic_key)
        if price <= 0:
            continue
        info = entry[lic_key]
        chosen.append({
            "name":  name,
            "bu":    info["bu"],
            "sheet": info["sheet"],
            "desc":  info["desc"],
            "qty":   1,
            "price": price,
        })
        if len(chosen) >= n:
            break
    return chosen


# ════════════════════════════════════════════════════════════════
#  테스트 케이스
# ════════════════════════════════════════════════════════════════
def run_tests():
    result = TestResult()
    df, cat = load_data()

    # ── TEST 1: 제품 1개 (구매) ─────────────────────────────────
    print("─" * 55)
    print("[TEST 1] 구매 제품 1개 → 출력 Excel에 정확히 1개")
    items_1 = pick_products(cat, df, "구매", n=1)
    if not items_1:
        result.fail("구매 제품을 카탈로그에서 찾지 못함 (SKIP)")
    else:
        buf = generate_excel(items_1[0]["bu"], items_1[0]["sheet"], DUMMY_INFO, items_1)
        verify_output(result, "TEST 1", items_1, buf)

    # ── TEST 2: 제품 2개 (구매, 같은 BU) ───────────────────────
    print("\n" + "─" * 55)
    print("[TEST 2] 구매 제품 2개 (같은 BU 템플릿) → 출력 Excel에 정확히 2개")
    items_2 = pick_products(cat, df, "구매", n=2)
    if len(items_2) < 2:
        result.fail("구매 제품 2개를 카탈로그에서 찾지 못함 (SKIP)")
    else:
        buf = generate_excel(items_2[0]["bu"], items_2[0]["sheet"], DUMMY_INFO, items_2)
        verify_output(result, "TEST 2", items_2, buf)

    # ── TEST 3: 제품 1개 (임대) ─────────────────────────────────
    print("\n" + "─" * 55)
    print("[TEST 3] 임대 제품 1개 → 출력 Excel에 정확히 1개")
    items_3 = pick_products(cat, df, "임대", n=1)
    if not items_3:
        result.fail("임대 제품을 카탈로그에서 찾지 못함 (SKIP)")
    else:
        buf = generate_excel(items_3[0]["bu"], items_3[0]["sheet"], DUMMY_INFO, items_3)
        verify_output(result, "TEST 3", items_3, buf)

    # ── TEST 4: 제품 2개 (임대) ─────────────────────────────────
    print("\n" + "─" * 55)
    print("[TEST 4] 임대 제품 2개 → 출력 Excel에 정확히 2개")
    items_4 = pick_products(cat, df, "임대", n=2)
    if len(items_4) < 2:
        result.fail("임대 제품 2개를 카탈로그에서 찾지 못함 (SKIP)")
    else:
        buf = generate_excel(items_4[0]["bu"], items_4[0]["sheet"], DUMMY_INFO, items_4)
        verify_output(result, "TEST 4", items_4, buf)

    # ── TEST 5: 수량 2개, 단가 직접 지정 ───────────────────────
    print("\n" + "─" * 55)
    print("[TEST 5] 수량=2 + 단가 커스텀 → 수량/단가 정확히 반영")
    items_5 = pick_products(cat, df, "구매", n=1)
    if not items_5:
        result.fail("제품을 찾지 못함 (SKIP)")
    else:
        items_5[0]["qty"]   = 3
        items_5[0]["price"] = 99_000_000
        buf = generate_excel(items_5[0]["bu"], items_5[0]["sheet"], DUMMY_INFO, items_5)
        verify_output(result, "TEST 5", items_5, buf)

    # ── TEST 6: 고객 정보 셀 기록 검증 ─────────────────────────
    print("\n" + "─" * 55)
    print("[TEST 6] 고객 정보가 Excel 셀에 올바르게 기록되는지")
    import openpyxl
    items_6 = pick_products(cat, df, "구매", n=1)
    if not items_6:
        result.fail("제품을 찾지 못함 (SKIP)")
    else:
        buf = generate_excel(items_6[0]["bu"], items_6[0]["sheet"], DUMMY_INFO, items_6)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        b8  = str(ws["B8"].value  or "")
        k9  = str(ws["K9"].value  or "")
        k10 = str(ws["K10"].value or "")
        k12 = str(ws["K12"].value or "")
        check(result, "B8 = 회사명 + 貴中",
              DUMMY_INFO["customer"] in b8,
              f"실제='{b8}'")
        check(result, "K9 = 견적일자",
              "2026" in k9,
              f"실제='{k9}'")
        check(result, "K10 = 담당자명",
              k10 == DUMMY_INFO["our_name"],
              f"실제='{k10}'")
        check(result, "K12 = 이메일",
              k12 == DUMMY_INFO["our_email"],
              f"실제='{k12}'")

    return result.summary()


if __name__ == "__main__":
    print("=" * 55)
    print("  Ansys 견적서 자동 검증 테스트")
    print("=" * 55)
    ok = run_tests()
    sys.exit(0 if ok else 1)
