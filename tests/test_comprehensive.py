#!/usr/bin/env python3
"""
다양한 인풋 조합 — 인풋/아웃풋 일치 검증
python -X utf8 tests/test_comprehensive.py
"""
import sys, os, io, re
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.excel_gen import (
    load_pricelist, build_product_catalog, generate_excel,
    get_price, read_products_from_excel, BU_FILES, DATA_DIR,
)
from datetime import date
import openpyxl

INFO = {
    "customer": "테스트", "contact": "홍길동", "tel": "02-0000",
    "our_name": "김담당", "our_tel": "010-0000", "our_email": "a@b.com",
    "issue_date": date(2026, 3, 4),
}

df  = load_pricelist()
cat = build_product_catalog()

PASS = FAIL = 0


def get_price_rows(bu, sheet):
    path = os.path.join(DATA_DIR, BU_FILES[bu])
    wb   = openpyxl.load_workbook(path, data_only=True)
    ws   = wb[sheet]
    rows = []
    for r in range(18, 120):
        b = str(ws.cell(r, 2).value or "").strip()
        i = ws.cell(r, 9).value
        if re.match(r"^\d+\.$", b) and i is not None and isinstance(i, (int, float)):
            rows.append(r)
    return rows


def make_item(name, lic, bu_override=None, sheet_override=None):
    entry = cat.get(name, {}).get(lic)
    if not entry:
        return None
    price = get_price(df, name, lic)
    if not price:
        return None
    return {
        "name":  name,
        "bu":    bu_override    or entry["bu"],
        "sheet": sheet_override or entry["sheet"],
        "qty":   1,
        "price": price,
        "desc":  entry["desc"],
    }


def test(label, items):
    global PASS, FAIL
    if not items or any(x is None for x in items):
        print(f"  [SKIP] {label}")
        return
    try:
        buf     = generate_excel(items[0]["bu"], items[0]["sheet"], INFO, items)
        out     = read_products_from_excel(buf)
        in_names  = [it["name"]  for it in items]
        out_names = [p["name"]   for p in out]

        mismatch = []
        if in_names != out_names:
            mismatch.append(f"이름 기대={in_names}  실제={out_names}")
        if len(items) == len(out):
            for it, p in zip(items, out):
                if p["qty"] != it["qty"]:
                    mismatch.append(f"수량 [{it['name'][:25]}] {it['qty']}→{p['qty']}")
                if p["price"] != it["price"]:
                    mismatch.append(f"단가 [{it['name'][:25]}] {it['price']:,}→{p['price']:,}")

        if not mismatch:
            PASS += 1
            print(f"  [PASS] {label}")
        else:
            FAIL += 1
            print(f"  [FAIL] {label}")
            for m in mismatch:
                print(f"         {m}")
    except Exception as e:
        FAIL += 1
        print(f"  [ERR]  {label}: {e}")


# ════════════════════════════════════════════════════════════
print("=" * 60)
print("  다양한 인풋 조합 검증")
print("=" * 60)

# ─ A. 단일 제품, BU별 대표 ─────────────────────────────────
print("\n[A] 단일 제품 — BU별 대표")
cases_a = [
    ("MBU",  "구매", "Ansys Mechanical Enterprise"),
    ("MBU",  "임대", "Ansys Discovery Pro"),
    ("FBU",  "구매", "Ansys CFD Enterprise"),
    ("FBU",  "임대", "Ansys CFD Premium"),
    ("EBU",  "구매", "Ansys Electronics Premium Maxwell"),
    ("EBU",  "임대", "Ansys Electronics Premium HFSS"),
    ("SBU",  "구매", "Ansys Twin Builder Enterprise"),
    ("CPBU+DBU+Matbu", "임대", "Ansys Minerva"),
]
for bu, lic, name in cases_a:
    it = make_item(name, lic)
    test(f"{bu}/{lic}  {name[:38]}", [it] if it else None)

# ─ B. 2개 — 복수 price_row 시트 ─────────────────────────────
print("\n[B] 2개 제품 — 복수 price_row 시트에 넣기")

# FBU/임대 Pre 시트 = price_rows [20, 41, 47]
it_cfd = make_item("Ansys CFD Premium", "임대")
it_hpc = make_item("Ansys HPC Pack",    "임대")
if it_cfd and it_hpc:
    pr = get_price_rows(it_cfd["bu"], it_cfd["sheet"])
    print(f"    (FBU/임대 Pre 시트 price_rows={pr})")
    # HPC도 같은 시트 사용 (아이템이 템플릿에 덮어씌워지는 케이스)
    it_hpc2 = {**it_hpc, "bu": it_cfd["bu"], "sheet": it_cfd["sheet"]}
    test("FBU/임대  CFD Premium + HPC Pack", [it_cfd, it_hpc2])

# MBU/구매 Pre 시트 = price_rows [20, 42]
it_mp = make_item("Ansys Mechanical Premium", "구매")
it_me = make_item("Ansys Mechanical Enterprise", "구매")
if it_mp and it_me:
    pr = get_price_rows(it_mp["bu"], it_mp["sheet"])
    print(f"    (MBU/구매 Pre 시트 price_rows={pr})")
    it_me2 = {**it_me, "bu": it_mp["bu"], "sheet": it_mp["sheet"]}
    test("MBU/구매  Mechanical Premium + Enterprise", [it_mp, it_me2])

# ─ C. 단일 price_row 시트에 2개 (오버플로우) ─────────────────
print("\n[C] 단일 price_row 시트에 2개 아이템 (오버플로우 케이스)")
it_me  = make_item("Ansys Mechanical Enterprise", "구매")  # MBU/Ent rows=[20]
it_mp2 = make_item("Ansys Mechanical Premium",    "구매")
if it_me and it_mp2:
    pr = get_price_rows(it_me["bu"], it_me["sheet"])
    print(f"    (MBU/구매 Ent 시트 price_rows={pr})")
    it_mp_over = {**it_mp2, "bu": it_me["bu"], "sheet": it_me["sheet"]}
    test("MBU/구매 Ent시트(rows=1)에 2개 아이템", [it_me, it_mp_over])

it_ent_ebu = make_item("Ansys Electronics Enterprise",        "구매")  # EBU rows=[20]
it_max_ebu = make_item("Ansys Electronics Premium Maxwell",   "구매")
if it_ent_ebu and it_max_ebu:
    pr = get_price_rows(it_ent_ebu["bu"], it_ent_ebu["sheet"])
    print(f"    (EBU/구매 Ent 시트 price_rows={pr})")
    it_max_over = {**it_max_ebu, "bu": it_ent_ebu["bu"], "sheet": it_ent_ebu["sheet"]}
    test("EBU/구매 Ent시트(rows=1)에 2개 아이템", [it_ent_ebu, it_max_over])

# ─ D. 수량 다양 ───────────────────────────────────────────────
print("\n[D] 수량 다양 (qty > 1)")
for qty in [2, 5, 10, 100]:
    it = make_item("Ansys Mechanical Enterprise", "구매")
    if it:
        test(f"qty={qty:>3}  Ansys Mechanical Enterprise", [{**it, "qty": qty}])

# ─ E. 단가 수동 변경 ──────────────────────────────────────────
print("\n[E] 단가 수동 변경")
it_base = make_item("Ansys CFD Premium", "임대")
if it_base:
    for price in [1, 1_000_000, 50_000_000, 150_000_000, 999_999_999]:
        it2  = {**it_base, "price": price}
        buf  = generate_excel(it2["bu"], it2["sheet"], INFO, [it2])
        out  = read_products_from_excel(buf)
        if out and out[0]["price"] == price:
            PASS += 1
            print(f"  [PASS] price={price:>13,} → 출력={out[0]['price']:>13,}")
        else:
            FAIL += 1
            actual = out[0]["price"] if out else "(없음)"
            print(f"  [FAIL] price={price:>13,} → 출력={actual}")

# ─ F. 스크린샷 재현: CFD Premium + Maxwell (다른 BU) ──────────
print("\n[F] 스크린샷 재현: CFD Premium(FBU/임대) + Maxwell(EBU/구매)")
it_cfd2 = make_item("Ansys CFD Premium",                    "임대")  # FBU/Pre
it_max2 = make_item("Ansys Electronics Premium Maxwell",    "구매")  # EBU/Max
if it_cfd2 and it_max2:
    pr = get_price_rows(it_cfd2["bu"], it_cfd2["sheet"])
    print(f"    사용 템플릿: {it_cfd2['bu']} / {it_cfd2['sheet']}  price_rows={pr}")
    test("CFD Premium + Maxwell (다른 BU)", [it_cfd2, it_max2])

# ─ G. 3개 이상 ────────────────────────────────────────────────
print("\n[G] 3개 이상 아이템")
if it_cfd2:
    items_3 = []
    for name in ["Ansys CFD Premium", "Ansys HPC Pack", "Ansys Discovery Pro"]:
        it = make_item(name, "임대") or make_item(name, "구매")
        if it:
            items_3.append({**it, "bu": it_cfd2["bu"], "sheet": it_cfd2["sheet"]})
    if len(items_3) == 3:
        pr = get_price_rows(it_cfd2["bu"], it_cfd2["sheet"])
        print(f"    FBU/Pre price_rows={pr}")
        test("FBU/임대 Pre시트(rows=3)에 3개 — 딱 맞음", items_3)

    # 4개 (rows=3 초과 — 오버플로우)
    it4 = make_item("Ansys CFD Enterprise", "임대") or make_item("Ansys CFD HPC Ultimate", "임대")
    if it4 and len(items_3) == 3:
        items_4 = items_3 + [{**it4, "bu": it_cfd2["bu"], "sheet": it_cfd2["sheet"]}]
        test("FBU/임대 Pre시트(rows=3)에 4개 — 1개 오버플로우", items_4)

# ─ H. qty+price 조합, 합계 검증 ───────────────────────────────
print("\n[H] qty × price = 합계(K열) 검증")
if it_cfd2:
    items_h = [
        {**it_cfd2, "qty": 3, "price": 12_670_000},
        {**it_cfd2, "name": "Ansys HPC Pack", "qty": 5, "price": 2_351_000,
         "bu": it_cfd2["bu"], "sheet": it_cfd2["sheet"]},
    ]
    try:
        buf = generate_excel(items_h[0]["bu"], items_h[0]["sheet"], INFO, items_h)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        ws = wb.active
        all_ok = True
        for it in items_h:
            for r in range(18, 120):
                c = str(ws.cell(r, 3).value or "")
                if c == it["name"]:
                    i_val = ws.cell(r, 9).value
                    j_val = ws.cell(r, 10).value
                    k_val = ws.cell(r, 11).value
                    expected_k = it["qty"] * it["price"]
                    ok_k = (k_val == expected_k)
                    if ok_k:
                        PASS += 1
                        print(f"  [PASS] {it['name'][:30]}  {it['qty']}×{it['price']:,}={expected_k:,}  K={k_val}")
                    else:
                        FAIL += 1
                        all_ok = False
                        print(f"  [FAIL] {it['name'][:30]}  기대K={expected_k:,}  실제K={k_val}")
                    break
    except Exception as e:
        FAIL += 1
        print(f"  [ERR]  합계 검증: {e}")

# ─ I. 고객 정보 셀 ────────────────────────────────────────────
print("\n[I] 고객 정보 셀 정확성")
it_cell = make_item("Ansys Mechanical Enterprise", "구매")
if it_cell:
    buf = generate_excel(it_cell["bu"], it_cell["sheet"], INFO, [it_cell])
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb.active
    checks = [
        ("B8  회사명+貴中", INFO["customer"] in str(ws["B8"].value or "")),
        ("K9  견적일자",    "2026" in str(ws["K9"].value or "")),
        ("K10 담당자명",    str(ws["K10"].value or "") == INFO["our_name"]),
        ("K12 이메일",      str(ws["K12"].value or "") == INFO["our_email"]),
    ]
    for label, ok in checks:
        if ok:
            PASS += 1
            print(f"  [PASS] {label}")
        else:
            FAIL += 1
            print(f"  [FAIL] {label}")

# ─ 최종 결과 ──────────────────────────────────────────────────
print()
print("=" * 60)
total = PASS + FAIL
print(f"결과: {PASS}/{total} 통과", end="")
if FAIL:
    print(f"  ({FAIL}개 실패)")
else:
    print("  — 모두 통과!")
print("=" * 60)
sys.exit(0 if FAIL == 0 else 1)
